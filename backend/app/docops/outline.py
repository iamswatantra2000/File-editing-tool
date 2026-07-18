"""
Build a compact structural outline of a document to feed into the agent prompt.
Keeps token usage low by summarising rather than dumping full text.
"""

from typing import Any, Dict, List

from docx import Document as DocxDocument
from openpyxl import Workbook


def docx_outline(doc: DocxDocument, max_para_chars: int = 120) -> Dict[str, Any]:
    """Return a JSON-serialisable outline of a .docx document."""
    headings: List[Dict] = []
    paragraphs: List[str] = []
    tables: List[Dict] = []

    for para in doc.paragraphs:
        style = para.style.name if para.style else ""
        text = para.text.strip()
        if not text:
            continue
        if style.startswith("Heading"):
            level = style.replace("Heading", "").strip()
            headings.append({"level": int(level) if level.isdigit() else 1, "text": text})
        else:
            paragraphs.append(text[:max_para_chars] + ("…" if len(text) > max_para_chars else ""))

    for i, table in enumerate(doc.tables):
        rows = len(table.rows)
        cols = len(table.columns)
        header = [cell.text.strip() for cell in table.rows[0].cells] if rows else []
        tables.append({"index": i, "rows": rows, "cols": cols, "header": header})

    return {
        "type": "docx",
        "headings": headings,
        "paragraphs": paragraphs[:20],  # cap to avoid prompt bloat
        "tables": tables,
    }


def xlsx_outline(wb: Workbook, max_rows: int = 5) -> Dict[str, Any]:
    """Return a JSON-serialisable outline of an .xlsx workbook."""
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        preview: List[List] = []
        for row in ws.iter_rows(min_row=1, max_row=max_rows, values_only=True):
            preview.append([str(v) if v is not None else "" for v in row])
        sheets.append({
            "name": name,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "preview": preview,
        })
    return {"type": "xlsx", "sheets": sheets}


def build_outline(session) -> Dict[str, Any]:
    """Dispatch to the right outline builder based on session type."""
    if session.ext == ".docx":
        return docx_outline(session.docx)
    return xlsx_outline(session.workbook)
