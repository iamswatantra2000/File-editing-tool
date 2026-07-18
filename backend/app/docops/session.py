"""
DocumentSession — opens a working copy and exposes the right ops object
based on file extension (.docx or .xlsx).
"""

from pathlib import Path
from typing import Union

from docx import Document as DocxDocument
from openpyxl import load_workbook


class DocumentSession:
    def __init__(self, working_path: Path):
        self.path = working_path
        self.ext = working_path.suffix.lower()

        if self.ext == ".docx":
            self._doc = DocxDocument(str(working_path))
            self._wb = None
        elif self.ext == ".xlsx":
            self._wb = load_workbook(str(working_path))
            self._doc = None
        else:
            raise ValueError(f"Unsupported file type: {self.ext}")

    # -- docx helpers ---------------------------------------------------------

    @property
    def docx(self) -> DocxDocument:
        if self._doc is None:
            raise RuntimeError("Not a .docx session")
        return self._doc

    def save_docx(self) -> None:
        self._doc.save(str(self.path))

    # -- xlsx helpers ---------------------------------------------------------

    @property
    def workbook(self):
        if self._wb is None:
            raise RuntimeError("Not a .xlsx session")
        return self._wb

    def save_xlsx(self) -> None:
        self._wb.save(str(self.path))

    # -- generic --------------------------------------------------------------

    def save(self) -> None:
        if self.ext == ".docx":
            self.save_docx()
        else:
            self.save_xlsx()

    def reload(self) -> None:
        """Re-open the file from disk (used during validation)."""
        if self.ext == ".docx":
            self._doc = DocxDocument(str(self.path))
        else:
            self._wb = load_workbook(str(self.path))
