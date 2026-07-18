"""
openpyxl operations for .xlsx files.
All functions accept a DocumentSession and mutate the working copy in place,
then call session.save_xlsx() before returning.
"""

from typing import Any, List, Optional

from app.docops.session import DocumentSession


def read_cell(session: DocumentSession, sheet_name: str, cell: str) -> str:
    """Return the string value of a single cell (e.g. 'A1')."""
    ws = session.workbook[sheet_name]
    value = ws[cell].value
    return str(value) if value is not None else ""


def write_cell(session: DocumentSession, sheet_name: str, cell: str, value: Any) -> str:
    """Write a value to a single cell."""
    ws = session.workbook[sheet_name]
    ws[cell] = value
    session.save_xlsx()
    return f"Set {sheet_name}!{cell} = {value!r}"


def write_range(
    session: DocumentSession,
    sheet_name: str,
    start_cell: str,
    rows: List[List[Any]],
) -> str:
    """Write a 2-D list of values starting at start_cell."""
    from openpyxl.utils.cell import coordinate_to_tuple
    ws = session.workbook[sheet_name]
    start_row, start_col = coordinate_to_tuple(start_cell)
    for r_offset, row in enumerate(rows):
        for c_offset, value in enumerate(row):
            ws.cell(row=start_row + r_offset, column=start_col + c_offset, value=value)
    session.save_xlsx()
    return f"Wrote {len(rows)} row(s) starting at {sheet_name}!{start_cell}"


def append_row(session: DocumentSession, sheet_name: str, values: List[Any]) -> str:
    """Append a row to the end of the used range."""
    ws = session.workbook[sheet_name]
    ws.append(values)
    session.save_xlsx()
    return f"Appended row to {sheet_name}: {values}"


def add_sheet(session: DocumentSession, sheet_name: str) -> str:
    """Add a new worksheet."""
    if sheet_name in session.workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' already exists")
    session.workbook.create_sheet(sheet_name)
    session.save_xlsx()
    return f"Added sheet '{sheet_name}'"


def delete_sheet(session: DocumentSession, sheet_name: str) -> str:
    """Delete a worksheet."""
    if sheet_name not in session.workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found")
    del session.workbook[sheet_name]
    session.save_xlsx()
    return f"Deleted sheet '{sheet_name}'"


def rename_sheet(session: DocumentSession, old_name: str, new_name: str) -> str:
    """Rename a worksheet."""
    if old_name not in session.workbook.sheetnames:
        raise ValueError(f"Sheet '{old_name}' not found")
    session.workbook[old_name].title = new_name
    session.save_xlsx()
    return f"Renamed sheet '{old_name}' → '{new_name}'"


def edit_table_cell(
    session: DocumentSession,
    sheet_name: str,
    row: int,
    col: int,
    value: Any,
) -> str:
    """Write to a cell by 1-based row/col indices."""
    ws = session.workbook[sheet_name]
    ws.cell(row=row, column=col, value=value)
    session.save_xlsx()
    return f"Set {sheet_name} row={row} col={col} = {value!r}"
